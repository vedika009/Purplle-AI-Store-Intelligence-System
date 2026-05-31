import openpyxl

wb = openpyxl.load_workbook('./data/raw/Brigade Road - Store layoutc5f5d56.xlsx')
for sheet in wb.sheetnames:
    print(f"Sheet: {sheet}")
    ws = wb[sheet]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        non_empty = {c_idx: val for c_idx, val in enumerate(row) if val is not None}
        if non_empty:
            print(f"Row {r_idx}: {non_empty}")
