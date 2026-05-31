import openpyxl

wb = openpyxl.load_workbook('./data/raw/Brigade Road - Store layoutc5f5d56.xlsx')
print("Sheets:", wb.sheetnames)
ws = wb.active
print("Dimensions:", ws.dimensions)
for r in range(1, 100):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 50)]
    non_empty = {c: val for c, val in enumerate(row_vals, 1) if val is not None}
    if non_empty:
        print(f"Row {r}: {non_empty}")
