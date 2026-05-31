import openpyxl

wb = openpyxl.load_workbook('./data/raw/Brigade Road - Store layoutc5f5d56.xlsx')
for sheet in wb.sheetnames:
    print(f"Sheet: {sheet}")
    ws = wb[sheet]
    for row in list(ws.iter_rows(values_only=True))[:15]:
        print(row)
